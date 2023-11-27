from openpyxl import load_workbook, chart


wb = load_workbook('./data/transactions.xlsx')
ws = wb['sheet_1']


for i in range(2, ws.max_row + 1):
    cell_value = float(ws[f'C{i}'].value.replace('€', '').replace(',', '.'))
    cell_value_discounted = round(cell_value * 0.90, 2)
    ws[f'D{i}'] = cell_value_discounted

# Chart
values = chart.Reference(ws, min_row=2, max_row=ws.max_row, min_col=4, max_col=4)
bar_chart = chart.BarChart()
bar_chart.add_data(values)
ws.add_chart(bar_chart, 'E2')

wb.save('./data/transactions_corrected.xlsx')